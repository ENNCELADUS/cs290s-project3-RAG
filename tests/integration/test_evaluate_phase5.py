from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from evaluate.cli import main as evaluate_main
from evaluate.export import SUBMISSION_COLUMNS


def test_evaluate_retrieve_exports_assignment_workbook_with_diagnostics(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"

    class FakeRetriever:
        def __init__(self, **kwargs: object) -> None:
            pass

        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls(**kwargs)

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            urls = {
                "dense": ["https://example.edu/noise", "https://example.edu/source"],
                "hybrid": ["https://example.edu/source"],
            }
            return [
                _Hit(rank=index, url=url, title=f"{mode}-{index}", score=1.0 / index)
                for index, url in enumerate(urls[mode], start=1)
            ]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    exit_code = evaluate_main(
        [
            "--questions",
            str(questions_path),
            "--output-dir",
            str(output_dir),
            "--runner",
            "retrieve",
            "--timestamp",
            "20260611T000000Z",
        ]
    )

    assert exit_code == 0
    workbook_path = output_dir / "results_before_after_20260611T000000Z.xlsx"
    summary_path = output_dir / "summary_20260611T000000Z.json"
    run_path = output_dir / "run_20260611T000000Z.jsonl"
    assert workbook_path.exists()
    assert summary_path.exists()
    assert run_path.exists()

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["submission", "diagnostics", "retrieval_metrics", "review_queue"]
    submission_rows = list(workbook["submission"].iter_rows(values_only=True))
    assert list(submission_rows[0]) == SUBMISSION_COLUMNS
    assert submission_rows[1][0] == "Where is the source?"
    assert submission_rows[1][2] is None
    assert submission_rows[1][3] is None
    assert submission_rows[1][4] is None
    assert submission_rows[1][5] is None

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary["modes"]["dense"]["source_hit@1"] == 0.0
    assert summary["modes"]["dense"]["source_hit@5"] == 1.0
    assert summary["modes"]["hybrid"]["mrr@5"] == 1.0


def test_evaluate_applies_review_decisions_to_submission_workbook(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    decisions_path = tmp_path / "review.csv"
    decisions_path.write_text(
        "id,mode,is_correct,review_note\nq1,dense,0,missed answer\nq1,hybrid,1,accepted\n", encoding="utf-8"
    )
    output_dir = tmp_path / "eval"

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "retrieve",
                "--review-decisions",
                str(decisions_path),
                "--timestamp",
                "20260611T010000Z",
            ]
        )
        == 0
    )

    workbook = load_workbook(output_dir / "results_before_after_20260611T010000Z.xlsx")
    submission_rows = list(workbook["submission"].iter_rows(values_only=True))
    assert submission_rows[1][4] == 0
    assert submission_rows[1][5] == 1


def test_evaluate_can_require_final_correctness_labels(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    with pytest.raises(ValueError, match="final correctness labels"):
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(tmp_path / "eval"),
                "--runner",
                "retrieve",
                "--require-final-labels",
                "--timestamp",
                "20260611T020000Z",
            ]
        )


def test_evaluate_final_labels_require_generated_answers(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    decisions_path = tmp_path / "review.csv"
    decisions_path.write_text("id,mode,is_correct\nq1,dense,1\nq1,hybrid,1\n", encoding="utf-8")

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    with pytest.raises(ValueError, match="final generated responses"):
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(tmp_path / "eval"),
                "--runner",
                "retrieve",
                "--review-decisions",
                str(decisions_path),
                "--require-final-labels",
                "--timestamp",
                "20260611T030000Z",
            ]
        )


def test_evaluate_answer_metrics_use_cited_sources_only(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

    class FakeAnswerer:
        def __init__(self, retriever: object, **kwargs: object) -> None:
            pass

        def answer(self, query: str, *, mode: str, top_k: int) -> _AnswerResult:
            return _AnswerResult(
                status="answered",
                answer="The answer cites only the first source [1].",
                sources=[
                    _Source(1, "https://example.edu/noise", "Noise"),
                    _Source(2, "https://example.edu/source", "Expected"),
                ],
                retrieval={"mode": mode, "hits": []},
            )

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)
    monkeypatch.setattr("evaluate.runner.RagAnswerer", FakeAnswerer)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "answer",
                "--model-path",
                str(tmp_path),
                "--timestamp",
                "20260611T040000Z",
            ]
        )
        == 0
    )

    summary = json.loads((output_dir / "summary_20260611T040000Z.json").read_text(encoding="utf-8"))
    assert summary["modes"]["dense"]["source_hit@5"] == 0.0
    run_records = [
        json.loads(line)
        for line in (output_dir / "run_20260611T040000Z.jsonl").read_text(encoding="utf-8").splitlines()
    ]
    assert run_records[0]["cited_source_urls"] == ["https://example.edu/noise"]
    assert run_records[0]["retrieved_source_urls"] == ["https://example.edu/noise", "https://example.edu/source"]


def test_evaluate_answer_passes_hybrid_knobs_to_generation_retrieval_only(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    calls: list[dict[str, object]] = []

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            calls.append({"mode": mode, "top_k": top_k, **kwargs})
            return []

        def contexts_for_hits(self, hits: list[object]) -> list[object]:
            return []

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(tmp_path / "eval"),
                "--runner",
                "answer",
                "--modes",
                "dense",
                "hybrid",
                "--model-path",
                str(tmp_path),
                "--device",
                "cpu",
                "--top-k",
                "5",
                "--sparse-top-k",
                "50",
                "--dense-top-k",
                "60",
                "--fused-top-k",
                "25",
                "--rerank-top-k",
                "20",
                "--rerank-preserve-top-k",
                "2",
                "--rrf-k",
                "70",
                "--sparse-weight",
                "0.8",
                "--dense-weight",
                "1.7",
                "--url-cap",
                "2",
                "--reranker-model",
                "/models/local-reranker",
                "--reranker-device",
                "cpu",
                "--expanded-query",
                "synthetic expansion",
                "--timestamp",
                "20260611T110000Z",
            ]
        )
        == 0
    )

    assert calls == [
        {"mode": "dense", "top_k": 5},
        {
            "mode": "hybrid",
            "top_k": 5,
            "sparse_top_k": 50,
            "dense_top_k": 60,
            "fused_top_k": 25,
            "rerank_top_k": 20,
            "rerank_preserve_top_k": 2,
            "rrf_k": 70,
            "sparse_weight": 0.8,
            "dense_weight": 1.7,
            "url_cap": 2,
            "reranker_model": "/models/local-reranker",
            "reranker_device": "cpu",
            "expanded_queries": ("synthetic expansion",),
        },
    ]


def test_evaluate_retrieve_can_save_diagnostic_hits_beyond_final_top_k(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"
    calls: list[dict[str, object]] = []

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            calls.append({"mode": mode, "top_k": top_k, **kwargs})
            return [
                _Hit(
                    rank=rank,
                    url="https://example.edu/source" if rank == 8 else f"https://example.edu/noise/{rank}",
                    title=f"{mode}-{rank}",
                    score=1.0 / rank,
                )
                for rank in range(1, top_k + 1)
            ]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "retrieve",
                "--diagnostic-depth",
                "25",
                "--timestamp",
                "20260611T070000Z",
            ]
        )
        == 0
    )

    summary = json.loads((output_dir / "summary_20260611T070000Z.json").read_text(encoding="utf-8"))
    assert summary["modes"]["dense"]["source_hit@5"] == 0.0
    assert summary["modes"]["hybrid"]["source_hit@5"] == 0.0

    run_records = [
        json.loads(line)
        for line in (output_dir / "run_20260611T070000Z.jsonl").read_text(encoding="utf-8").splitlines()
    ]
    assert [record["top_titles"] for record in run_records] == [
        ["dense-1", "dense-2", "dense-3", "dense-4", "dense-5"],
        ["hybrid-1", "hybrid-2", "hybrid-3", "hybrid-4", "hybrid-5"],
    ]
    assert [len(record["diagnostic_hits"]) for record in run_records] == [25, 25]
    assert run_records[0]["diagnostic_hits"][7]["url"] == "https://example.edu/source"
    hybrid_diagnostic_call = calls[3]
    assert hybrid_diagnostic_call["mode"] == "hybrid"
    assert hybrid_diagnostic_call["top_k"] == 25
    assert hybrid_diagnostic_call["sparse_top_k"] == 25
    assert hybrid_diagnostic_call["dense_top_k"] == 25
    assert hybrid_diagnostic_call["fused_top_k"] == 25


def test_evaluate_retrieve_passes_hybrid_knobs_only_to_hybrid(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"
    calls: list[dict[str, object]] = []

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            calls.append({"mode": mode, "top_k": top_k, **kwargs})
            return [
                _Hit(rank=rank, url="https://example.edu/source", title=f"{mode}-{rank}", score=1.0 / rank)
                for rank in range(1, top_k + 1)
            ]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "retrieve",
                "--modes",
                "dense",
                "hybrid",
                "--top-k",
                "5",
                "--diagnostic-depth",
                "25",
                "--sparse-top-k",
                "50",
                "--dense-top-k",
                "50",
                "--fused-top-k",
                "50",
                "--rerank-top-k",
                "50",
                "--rerank-preserve-top-k",
                "2",
                "--rrf-k",
                "70",
                "--sparse-weight",
                "0.8",
                "--dense-weight",
                "1.7",
                "--url-cap",
                "2",
                "--expanded-query",
                "synthetic expansion",
                "--timestamp",
                "20260611T090000Z",
            ]
        )
        == 0
    )

    assert len(calls) == 4
    dense_call, dense_diagnostic_call, hybrid_call, hybrid_diagnostic_call = calls
    hybrid_keys = {
        "sparse_top_k",
        "dense_top_k",
        "fused_top_k",
        "rerank_top_k",
        "rerank_preserve_top_k",
        "rrf_k",
        "sparse_weight",
        "dense_weight",
        "url_cap",
        "expanded_queries",
    }
    assert dense_call == {"mode": "dense", "top_k": 5}
    assert dense_diagnostic_call == {"mode": "dense", "top_k": 25}
    assert hybrid_keys.isdisjoint(dense_call)
    assert hybrid_keys.isdisjoint(dense_diagnostic_call)
    assert hybrid_call == {
        "mode": "hybrid",
        "top_k": 5,
        "sparse_top_k": 50,
        "dense_top_k": 50,
        "fused_top_k": 50,
        "rerank_top_k": 50,
        "rerank_preserve_top_k": 2,
        "rrf_k": 70,
        "sparse_weight": 0.8,
        "dense_weight": 1.7,
        "url_cap": 2,
        "expanded_queries": ("synthetic expansion",),
    }
    assert hybrid_diagnostic_call == {
        "mode": "hybrid",
        "top_k": 25,
        "sparse_top_k": 50,
        "dense_top_k": 50,
        "fused_top_k": 50,
        "rerank_top_k": 50,
        "rerank_preserve_top_k": 2,
        "rrf_k": 70,
        "sparse_weight": 0.8,
        "dense_weight": 1.7,
        "url_cap": 2,
        "expanded_queries": ("synthetic expansion",),
    }


def test_evaluate_retrieve_passes_reranker_model_and_device_only_to_hybrid(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"
    calls: list[dict[str, object]] = []

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            calls.append({"mode": mode, "top_k": top_k, **kwargs})
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "retrieve",
                "--reranker-model",
                "/models/local-reranker",
                "--reranker-device",
                "cuda",
                "--diagnostic-depth",
                "10",
                "--timestamp",
                "20260611T080000Z",
            ]
        )
        == 0
    )

    assert calls[0]["mode"] == "dense"
    assert "reranker_model" not in calls[0]
    assert "reranker_device" not in calls[0]
    assert calls[1]["mode"] == "dense"
    assert "reranker_model" not in calls[1]
    assert "reranker_device" not in calls[1]
    assert calls[2]["mode"] == "hybrid"
    assert calls[2]["reranker_model"] == "/models/local-reranker"
    assert calls[2]["reranker_device"] == "cuda"
    assert calls[3]["mode"] == "hybrid"
    assert calls[3]["reranker_model"] == "/models/local-reranker"
    assert calls[3]["reranker_device"] == "cuda"


def test_evaluate_retrieve_passes_expanded_queries_jsonl_only_to_hybrid(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    expansions_path = tmp_path / "expanded_queries.jsonl"
    expansions_path.write_text(
        json.dumps({"id": "q1", "expanded_queries": ["local Qwen pseudo answer", "alternate course phrasing"]}) + "\n",
        encoding="utf-8",
    )
    calls: list[dict[str, object]] = []

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            calls.append({"mode": mode, "top_k": top_k, **kwargs})
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(tmp_path / "eval"),
                "--runner",
                "retrieve",
                "--expanded-queries-jsonl",
                str(expansions_path),
                "--timestamp",
                "20260611T100000Z",
            ]
        )
        == 0
    )

    assert calls[0] == {"mode": "dense", "top_k": 5}
    assert calls[1] == {
        "mode": "hybrid",
        "top_k": 5,
        "expanded_queries": ("local Qwen pseudo answer", "alternate course phrasing"),
    }


def test_evaluate_both_runner_emits_retrieval_and_answer_records(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)
    output_dir = tmp_path / "eval"

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            return [_Hit(rank=1, url="https://example.edu/source", title=mode, score=1.0)]

    class FakeAnswerer:
        def __init__(self, retriever: object, **kwargs: object) -> None:
            pass

        def answer(self, query: str, *, mode: str, top_k: int) -> _AnswerResult:
            return _AnswerResult(
                status="answered",
                answer="At the expected source. [1]",
                sources=[_Source(1, "https://example.edu/source", "Expected")],
                retrieval={"mode": mode, "hits": []},
            )

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)
    monkeypatch.setattr("evaluate.runner.RagAnswerer", FakeAnswerer)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(output_dir),
                "--runner",
                "both",
                "--model-path",
                str(tmp_path),
                "--timestamp",
                "20260611T050000Z",
            ]
        )
        == 0
    )

    run_records = [
        json.loads(line)
        for line in (output_dir / "run_20260611T050000Z.jsonl").read_text(encoding="utf-8").splitlines()
    ]
    assert [(record["runner"], record["mode"]) for record in run_records] == [
        ("retrieve", "dense"),
        ("answer", "dense"),
        ("retrieve", "hybrid"),
        ("answer", "hybrid"),
    ]


def test_evaluate_returns_nonzero_when_any_record_errors(tmp_path: Path, monkeypatch) -> None:
    questions_path = tmp_path / "questions.csv"
    _write_questions(questions_path)

    class FakeRetriever:
        @classmethod
        def from_paths(cls, **kwargs: object) -> FakeRetriever:
            return cls()

        def retrieve(self, query: str, *, mode: str, top_k: int, **kwargs: object) -> list[object]:
            raise RuntimeError("broken retrieval")

    monkeypatch.setattr("evaluate.runner.Retriever", FakeRetriever)

    assert (
        evaluate_main(
            [
                "--questions",
                str(questions_path),
                "--output-dir",
                str(tmp_path / "eval"),
                "--runner",
                "retrieve",
                "--timestamp",
                "20260611T060000Z",
            ]
        )
        == 1
    )


class _Hit:
    def __init__(self, *, rank: int, url: str, title: str, score: float) -> None:
        self.rank = rank
        self.url = url
        self.title = title
        self.score = score
        self.chunk_id = rank
        self.document_id = rank
        self.snippet = f"snippet {rank}"


class _Source:
    def __init__(self, source_id: int, url: str, title: str) -> None:
        self.source_id = source_id
        self.url = url
        self.title = title


class _AnswerResult:
    def __init__(self, *, status: str, answer: str, sources: list[_Source], retrieval: dict[str, object]) -> None:
        self.status = status
        self.answer = answer
        self.sources = sources
        self.retrieval = retrieval


def _write_questions(path: Path) -> None:
    rows = [
        {
            "id": "q1",
            "category": "Factual",
            "language": "en",
            "query": "Where is the source?",
            "gt_answer": "At the expected source.",
            "primary_source_url": "https://example.edu/source",
            "acceptable_source_urls": json.dumps(["https://example.edu/source"]),
            "evidence_snippet": "expected source",
            "required_facts": json.dumps(["expected source"]),
            "acceptable_answers": json.dumps(["At the expected source."]),
            "forbidden_facts": json.dumps([]),
            "grading_notes": "test row",
            "judge_type": "exact_or_alias_match",
            "complexity": "Low",
            "sys_resp_before_opt": "",
            "sys_resp_after_opt": "",
            "is_correct_before_opt": "",
            "is_correct_after_opt": "",
            "cited_expected_source_hit_before_opt": "",
            "cited_expected_source_hit_after_opt": "",
        }
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
