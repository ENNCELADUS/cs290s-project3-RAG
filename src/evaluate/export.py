from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook

SUBMISSION_COLUMNS = [
    "query",
    "gt_answer",
    "sys_resp_before_opt",
    "sys_resp_after_opt",
    "is_correct_before_opt",
    "is_correct_after_opt",
]


def write_outputs(
    records: list[dict[str, Any]],
    output_dir: Path,
    *,
    timestamp: str,
    review_decisions: dict[tuple[str, str], int] | None = None,
    require_final_labels: bool = False,
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    run_path = output_dir / f"run_{timestamp}.jsonl"
    summary_path = output_dir / f"summary_{timestamp}.json"
    review_path = output_dir / f"review_queue_{timestamp}.csv"
    gap_path = output_dir / f"gap_notes_{timestamp}.md"
    workbook_path = output_dir / f"results_before_after_{timestamp}.xlsx"
    submission_rows = _submission_rows(records, review_decisions=review_decisions or {})
    if require_final_labels:
        missing_labels = [
            row["query"]
            for row in submission_rows
            if row["is_correct_before_opt"] is None or row["is_correct_after_opt"] is None
        ]
        if missing_labels:
            raise ValueError(f"final correctness labels are missing for {len(missing_labels)} questions")
        missing_responses = [
            row["query"] for row in submission_rows if not row["sys_resp_before_opt"] or not row["sys_resp_after_opt"]
        ]
        if missing_responses:
            raise ValueError(f"final generated responses are missing for {len(missing_responses)} questions")

    _write_jsonl(run_path, records)
    summary = summarize_records(records)
    _write_json(summary_path, summary)
    _write_review_queue(review_path, records)
    _write_gap_notes(gap_path, records)
    _write_workbook(workbook_path, records, summary, submission_rows=submission_rows)
    return {
        "run": run_path,
        "summary": summary_path,
        "review_queue": review_path,
        "gap_notes": gap_path,
        "workbook": workbook_path,
    }


def summarize_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    by_mode: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_mode[str(record["mode"])].append(record)
    summary: dict[str, Any] = {"total_records": len(records), "modes": {}}
    for mode, mode_records in sorted(by_mode.items()):
        ok_records = [record for record in mode_records if record["status"] == "ok"]
        latencies = [float(record["latency_s"]) for record in ok_records]
        metric_names = sorted({name for record in ok_records for name in dict(record.get("metrics", {}))})
        summary["modes"][mode] = {
            "records": len(mode_records),
            "ok": len(ok_records),
            "errors": len(mode_records) - len(ok_records),
            "avg_latency_s": round(mean(latencies), 6) if latencies else None,
            **{
                name: round(mean(float(dict(record.get("metrics", {})).get(name, 0.0)) for record in ok_records), 6)
                if ok_records
                else 0.0
                for name in metric_names
            },
            "correct": _judge_count(ok_records, "correct"),
            "incorrect": _judge_count(ok_records, "incorrect"),
            "manual_review": _judge_count(ok_records, "manual_review"),
            "evidence_insufficient": _judge_count(ok_records, "evidence_insufficient"),
        }
    return summary


def _write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, sort_keys=True, indent=2)
        handle.write("\n")


def _write_review_queue(path: Path, records: list[dict[str, Any]]) -> None:
    rows = [record for record in records if dict(record.get("judge", {})).get("status") == "manual_review"]
    fieldnames = ["id", "mode", "query", "gt_answer", "answer", "judge_reason", "observed_source_urls"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in rows:
            writer.writerow(
                {
                    "id": record["id"],
                    "mode": record["mode"],
                    "query": record["query"],
                    "gt_answer": record["gt_answer"],
                    "answer": record.get("answer") or "",
                    "judge_reason": dict(record.get("judge", {})).get("reason", ""),
                    "observed_source_urls": json.dumps(record.get("observed_source_urls", []), ensure_ascii=False),
                }
            )


def _write_gap_notes(path: Path, records: list[dict[str, Any]]) -> None:
    lines = ["# Evaluation Gap Notes", ""]
    for record in records:
        metrics = dict(record.get("metrics", {}))
        judge = dict(record.get("judge", {}))
        if record["status"] != "ok" or metrics.get("source_hit@5") != 1.0 or judge.get("status") != "correct":
            lines.append(f"- `{record['id']}` `{record['mode']}`: {_failure_reason(record)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_workbook(
    path: Path,
    records: list[dict[str, Any]],
    summary: dict[str, Any],
    *,
    submission_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    submission = workbook.active
    submission.title = "submission"
    submission.append(SUBMISSION_COLUMNS)
    for row in submission_rows:
        submission.append([row.get(column) for column in SUBMISSION_COLUMNS])

    diagnostics = workbook.create_sheet("diagnostics")
    diagnostic_columns = [
        "id",
        "mode",
        "query",
        "status",
        "answer_status",
        "answer",
        "judge_status",
        "judge_reason",
        "source_hit@5",
        "retrieved_expected_source_hit@5",
        "retrieved_expected_source_recall@5",
        "all_retrieved_expected_sources_hit@5",
        "cited_expected_source_hit@5",
        "answer_synthesis_miss",
        "generation_path",
        "generation_rejection_reason",
        "fallback_source_rank",
        "answer_context_order",
        "mrr@5",
        "latency_s",
        "observed_source_urls",
    ]
    diagnostics.append(diagnostic_columns)
    for record in records:
        metrics = dict(record.get("metrics", {}))
        judge = dict(record.get("judge", {}))
        diagnostics.append(
            [
                record.get("id"),
                record.get("mode"),
                record.get("query"),
                record.get("status"),
                record.get("answer_status"),
                record.get("answer"),
                judge.get("status"),
                judge.get("reason"),
                metrics.get("source_hit@5"),
                metrics.get("retrieved_expected_source_hit@5"),
                metrics.get("retrieved_expected_source_recall@5"),
                metrics.get("all_retrieved_expected_sources_hit@5"),
                metrics.get("cited_expected_source_hit@5"),
                metrics.get("answer_synthesis_miss"),
                record.get("generation_path"),
                record.get("generation_rejection_reason"),
                record.get("fallback_source_rank"),
                json.dumps(record.get("answer_context_order", []), ensure_ascii=False),
                metrics.get("mrr@5"),
                record.get("latency_s"),
                json.dumps(record.get("observed_source_urls", []), ensure_ascii=False),
            ]
        )

    retrieval_metrics = workbook.create_sheet("retrieval_metrics")
    metric_columns = [
        "mode",
        "records",
        "ok",
        "errors",
        "source_hit@1",
        "source_hit@5",
        "retrieved_expected_source_hit@5",
        "retrieved_expected_source_recall@5",
        "all_retrieved_expected_sources_hit@5",
        "cited_expected_source_hit@5",
        "answer_synthesis_miss",
        "source_recall@5",
        "mrr@5",
        "ndcg@5",
        "precision@5",
        "avg_latency_s",
    ]
    retrieval_metrics.append(metric_columns)
    for mode, payload in dict(summary.get("modes", {})).items():
        retrieval_metrics.append([mode, *[payload.get(column) for column in metric_columns[1:]]])

    review_queue = workbook.create_sheet("review_queue")
    review_queue.append(["id", "mode", "query", "gt_answer", "answer", "judge_reason"])
    for record in records:
        judge = dict(record.get("judge", {}))
        if judge.get("status") == "manual_review":
            review_queue.append(
                [
                    record.get("id"),
                    record.get("mode"),
                    record.get("query"),
                    record.get("gt_answer"),
                    record.get("answer"),
                    judge.get("reason"),
                ]
            )

    workbook.save(path)


def _submission_rows(
    records: list[dict[str, Any]], *, review_decisions: dict[tuple[str, str], int]
) -> list[dict[str, Any]]:
    by_question: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for record in records:
        by_question[str(record["id"])][str(record["mode"])] = record
    rows: list[dict[str, Any]] = []
    for question_id in sorted(by_question):
        records_by_mode = by_question[question_id]
        before = records_by_mode.get("dense")
        after = records_by_mode.get("hybrid")
        reference = before or after or next(iter(records_by_mode.values()))
        rows.append(
            {
                "query": reference.get("query"),
                "gt_answer": reference.get("gt_answer"),
                "sys_resp_before_opt": before.get("answer") if before else None,
                "sys_resp_after_opt": after.get("answer") if after else None,
                "is_correct_before_opt": _correctness_value(before, review_decisions) if before else None,
                "is_correct_after_opt": _correctness_value(after, review_decisions) if after else None,
            }
        )
    return rows


def _judge_count(records: list[dict[str, Any]], status: str) -> int:
    return sum(1 for record in records if dict(record.get("judge", {})).get("status") == status)


def _correctness_value(record: dict[str, Any], review_decisions: dict[tuple[str, str], int]) -> int | None:
    key = (str(record["id"]), str(record["mode"]))
    if key in review_decisions:
        return review_decisions[key]
    return dict(record.get("judge", {})).get("is_correct")


def _failure_reason(record: dict[str, Any]) -> str:
    if record["status"] != "ok":
        return str(record.get("error") or "run error")
    judge = dict(record.get("judge", {}))
    if dict(record.get("metrics", {})).get("answer_synthesis_miss") == 1.0:
        return "retrieved expected source but answer synthesis missed citation or abstained"
    if judge.get("status") == "evidence_insufficient":
        return str(judge.get("reason") or "evidence insufficient answer")
    if dict(record.get("metrics", {})).get("source_hit@5") != 1.0:
        return "expected official source not found in top 5"
    if judge.get("status") == "manual_review":
        return str(judge.get("reason") or "manual review required")
    if judge.get("status") == "incorrect":
        return str(judge.get("reason") or "incorrect answer")
    return "weak case"
